"""
Módulo importador de productos desde archivo Excel
Autor: Sistema de importación DeliciasNaturales
Fecha: Agosto 2025

Este módulo permite importar productos masivamente desde un archivo Excel
con validación de datos, vista previa e importación progresiva.
"""

from flask import Blueprint, request, jsonify
import openpyxl
import os
import tempfile
from datetime import datetime
import traceback
import time
from typing import Dict, List, Tuple, Optional

# Importaciones del proyecto
from models import db, Producto, Proveedor, Categoria, Marca
from config.database import get_db

# Crear el blueprint
importador_bp = Blueprint('importador', __name__)

class ImportadorProductos:
    """Clase principal para manejar la importación de productos desde Excel"""
    
    def __init__(self):
        self.productos_validados = []
        self.productos_excluidos = []
        self.archivo_temporal = None
        self.total_filas = 0
        self.progreso_actual = 0
        
    def procesar_archivo_excel(self, archivo) -> Dict:
        """
        Procesa un archivo Excel y retorna vista previa de productos
        
        Args:
            archivo: Archivo Excel subido
            
        Returns:
            Dict con productos válidos e inválidos para vista previa
        """
        try:
            print("🔍 [PROCESAR EXCEL] Iniciando procesamiento...")
            print(f"🔍 [DEBUG] Archivo: {archivo.filename}")
            print(f"🔍 [DEBUG] Content-Type: {archivo.content_type}")
            
            # Guardar archivo temporalmente
            print("📁 [DEBUG] Creando archivo temporal...")
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            archivo.save(temp_file.name)
            self.archivo_temporal = temp_file.name
            print(f"📁 [DEBUG] Archivo temporal guardado en: {temp_file.name}")
            
            # Verificar que el archivo se guardó correctamente
            if not os.path.exists(temp_file.name):
                print("❌ [ERROR] El archivo temporal no se pudo crear")
                return {
                    'error': True,
                    'mensaje': 'Error al guardar archivo temporal'
                }
            
            tamaño_archivo = os.path.getsize(temp_file.name)
            print(f"📊 [DEBUG] Tamaño del archivo: {tamaño_archivo} bytes")
            
            # Abrir el archivo Excel
            print("📊 [DEBUG] Abriendo archivo Excel con openpyxl...")
            # data_only=True hace que lea valores calculados en lugar de fórmulas
            workbook = openpyxl.load_workbook(temp_file.name, data_only=True)
            hoja = workbook.active
            print(f"📊 [DEBUG] Hoja activa: {hoja.title}")
            print(f"📊 [DEBUG] Dimensiones de la hoja: {hoja.max_row} filas x {hoja.max_column} columnas")
            
            # Resetear listas
            self.productos_validados = []
            self.productos_excluidos = []
            
            # Configurar límite de procesamiento
            max_filas_procesar = 1000  # Limitar para evitar timeout
            
            print(f"📊 [DEBUG] Procesando filas desde la fila 2...")
            print(f"📊 [DEBUG] Límite de procesamiento: {max_filas_procesar} filas")
            
            # Obtener todas las filas con datos (desde fila 2) - solo las primeras 1000 para evitar timeout
            filas_datos = []
            contador_filas = 0
            
            print(f"🔍 [DEBUG] Iniciando lectura de filas...")
            for index, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True)):
                if any(fila):  # Solo agregar filas que tienen datos
                    filas_datos.append(fila)
                    contador_filas += 1
                    
                    # Log cada 100 filas leídas
                    if contador_filas % 100 == 0:
                        print(f"📖 [LECTURA] Leídas {contador_filas} filas con datos...")
                    
                    # Limitar número de filas procesadas
                    if contador_filas >= max_filas_procesar:
                        print(f"⚠️ [WARNING] Archivo muy grande. Procesando solo las primeras {max_filas_procesar} filas")
                        break
                
                # También contar filas vacías para mostrar progreso total
                if (index + 1) % 200 == 0:
                    print(f"🔍 [ESCANEO] Escaneadas {index + 1} filas del Excel (encontradas {contador_filas} con datos)...")
            
            self.total_filas = len(filas_datos)
            print(f"📊 [DEBUG] Total de filas a procesar: {self.total_filas}")
            print(f"📊 [DEBUG] Primeras 3 filas de datos:")
            for i, fila in enumerate(filas_datos[:3]):
                print(f"  Fila {i+2}: {[str(cell)[:30] + '...' if cell and len(str(cell)) > 30 else cell for cell in fila[:10]]}")
            
            if self.total_filas == 0:
                return {
                    'error': True,
                    'mensaje': 'El archivo Excel está vacío o no tiene datos válidos en las primeras columnas'
                }
            
            # Procesar cada fila
            print(f"🔄 [DEBUG] Iniciando procesamiento de productos...")
            productos_procesados = 0
            
            for numero_fila, fila in enumerate(filas_datos, start=2):
                productos_procesados += 1
                
                # Log detallado cada 50 productos y siempre los primeros 10
                if productos_procesados <= 10 or productos_procesados % 50 == 0:
                    print(f"🔄 [DEBUG] Procesando fila {numero_fila} ({productos_procesados}/{self.total_filas})")
                    print(f"📊 [FILA] Datos: {[str(cell)[:50] + '...' if cell and len(str(cell)) > 50 else cell for cell in fila[:10]]}")
                
                resultado_producto = self._procesar_fila_producto(fila, numero_fila)
                
                if resultado_producto['valido']:
                    self.productos_validados.append(resultado_producto['producto'])
                    if productos_procesados <= 5:  # Log detallado de los primeros 5 productos válidos
                        print(f"✅ [PRODUCTO VÁLIDO] {resultado_producto['producto'].get('nombre', 'Sin nombre')}")
                else:
                    self.productos_excluidos.append(resultado_producto)
                    if productos_procesados <= 5:  # Log detallado de los primeros 5 productos excluidos
                        print(f"❌ [PRODUCTO EXCLUIDO] Fila {numero_fila}: {resultado_producto.get('motivo', 'Sin motivo')}")
                
                # Progreso cada 100 productos
                if productos_procesados % 100 == 0:
                    porcentaje = round((productos_procesados / self.total_filas) * 100, 1)
                    print(f"📈 [PROGRESO] {productos_procesados}/{self.total_filas} ({porcentaje}%) - Válidos: {len(self.productos_validados)}, Excluidos: {len(self.productos_excluidos)}")
            
            workbook.close()
            print(f"✅ [DEBUG] Procesamiento completado")
            print(f"📊 [STATS] Productos válidos: {len(self.productos_validados)}")
            print(f"📊 [STATS] Productos excluidos: {len(self.productos_excluidos)}")
            
            return {
                'error': False,
                'productos_validados': self.productos_validados,
                'productos_excluidos': self.productos_excluidos,
                'total_procesadas': self.total_filas,
                'total_validas': len(self.productos_validados),
                'total_invalidas': len(self.productos_excluidos)
            }
            
        except Exception as e:
            return {
                'error': True,
                'mensaje': f'Error procesando archivo Excel: {str(e)}',
                'detalle': traceback.format_exc()
            }
    
    def _procesar_fila_producto(self, fila: Tuple, numero_fila: int) -> Dict:
        """
        Procesa una fila individual del Excel y valida los datos
        
        Args:
            fila: Tupla con los valores de la fila
            numero_fila: Número de fila en el Excel
            
        Returns:
            Dict con información del producto procesado
        """
        try:
            # Mapear columnas según especificación
            # A2: Producto, C2: Proveedor, E2: Precio de Costo, F2: % ganancia,
            # G2: Precio Ganancia Paquete, H2: Precio por Unidad/100gr, I2: Descripción
            
            # Log detallado solo para las primeras 5 filas
            debug_detallado = numero_fila <= 7  # Primeras 5 filas de datos (fila 2-6)
            
            if debug_detallado:
                print(f"🔍 [FILA {numero_fila}] Analizando columnas:")
                print(f"  A (Producto): '{self._obtener_valor_celda(fila, 0)}'")
                print(f"  C (Proveedor): '{self._obtener_valor_celda(fila, 2)}'")
                print(f"  E (Precio Costo): '{self._obtener_valor_celda(fila, 4)}'")
                print(f"  F (% Ganancia): '{self._obtener_valor_celda(fila, 5)}'")
                print(f"  G (Precio Paquete): '{self._obtener_valor_celda(fila, 6)}'")
                print(f"  H (Precio Unidad): '{self._obtener_valor_celda(fila, 7)}'")
                print(f"  I (Descripción): '{self._obtener_valor_celda(fila, 8)}'")
            
            nombre_producto = self._obtener_valor_celda(fila, 0)  # Columna A
            proveedor_nombre = self._obtener_valor_celda(fila, 2)  # Columna C
            precio_costo = self._obtener_valor_numerico(fila, 4)  # Columna E
            porcentaje_ganancia = self._obtener_valor_numerico(fila, 5)  # Columna F
            precio_ganancia_paquete = self._obtener_valor_numerico(fila, 6)  # Columna G
            precio_por_unidad_100gr = self._obtener_valor_numerico(fila, 7)  # Columna H
            descripcion = self._obtener_valor_celda(fila, 8)  # Columna I
            
            if debug_detallado:
                print(f"🔍 [VALIDACIÓN FILA {numero_fila}] Valores procesados:")
                print(f"  Nombre: '{nombre_producto}'")
                print(f"  Proveedor: '{proveedor_nombre}'")
                print(f"  Precio Costo: {precio_costo}")
                print(f"  % Ganancia: {porcentaje_ganancia}")
                print(f"  Precio Paquete: {precio_ganancia_paquete}")
                print(f"  Precio Unidad: {precio_por_unidad_100gr}")
                print(f"  Descripción: '{descripcion}'")
            
            # Lista de errores de validación
            errores = []
            
            # Validaciones obligatorias
            if not nombre_producto:
                errores.append("Nombre del producto es obligatorio")
            
            if not proveedor_nombre:
                errores.append("Nombre del proveedor es obligatorio")
                
            if precio_costo is None:
                errores.append("Precio de costo es obligatorio")
            elif precio_costo <= 0:
                errores.append("Precio de costo debe ser mayor a 0")
                
            if porcentaje_ganancia is None:
                errores.append("Porcentaje de ganancia es obligatorio")
            elif porcentaje_ganancia < 0:
                errores.append("Porcentaje de ganancia no puede ser negativo")
                
            if precio_ganancia_paquete is None:
                errores.append("Precio de ganancia del paquete es obligatorio")
            elif precio_ganancia_paquete <= 0:
                errores.append("Precio de ganancia del paquete debe ser mayor a 0")
            
            # Determinar tipo de producto según especificación
            tipo_producto = 2 if precio_por_unidad_100gr is not None else 1
            unidad = "gramos" if tipo_producto == 2 else "unidad"
            
            # Si hay errores, retornar producto inválido
            if errores:
                if debug_detallado:
                    print(f"❌ [FILA {numero_fila}] PRODUCTO EXCLUIDO: {errores}")
                
                return {
                    'valido': False,
                    'fila': numero_fila,
                    'nombre': nombre_producto or "Sin nombre",
                    'errores': errores,
                    'datos_originales': {
                        'nombre': nombre_producto,
                        'proveedor': proveedor_nombre,
                        'precio_costo': precio_costo,
                        'porcentaje_ganancia': porcentaje_ganancia,
                        'precio_ganancia_paquete': precio_ganancia_paquete,
                        'precio_por_unidad_100gr': precio_por_unidad_100gr,
                        'descripcion': descripcion
                    }
                }
            
            # Crear objeto producto válido
            producto_datos = {
                'fila': numero_fila,
                'nombre': nombre_producto.strip(),
                'proveedor': proveedor_nombre.strip(),  # Cambiar a 'proveedor' para consistencia
                'precio_costo': float(precio_costo),
                'porcentaje_ganancia': float(porcentaje_ganancia / 100),  # Convertir a decimal
                'precio_ganancia_paquete': float(precio_ganancia_paquete),
                'precio_por_unidad_100gr': float(precio_por_unidad_100gr) if precio_por_unidad_100gr else None,
                'descripcion': descripcion.strip() if descripcion else "",
                'tipo': tipo_producto,  # Cambiar a 'tipo' para consistencia
                'unidad': unidad,
                'precio_calculado': float(precio_costo) * (1 + float(porcentaje_ganancia / 100))
            }
            
            if debug_detallado:
                print(f"✅ [FILA {numero_fila}] PRODUCTO VÁLIDO: '{nombre_producto}' - Tipo: {tipo_producto} ({unidad})")
            
            return {
                'valido': True,
                'producto': producto_datos
            }
            
        except Exception as e:
            return {
                'valido': False,
                'fila': numero_fila,
                'nombre': "Error de procesamiento",
                'errores': [f"Error inesperado: {str(e)}"],
                'datos_originales': {}
            }
    
    def _obtener_valor_celda(self, fila: Tuple, indice: int) -> Optional[str]:
        """Obtiene valor de celda como string, manejando valores None"""
        try:
            valor = fila[indice] if indice < len(fila) else None
            return str(valor).strip() if valor is not None else None
        except:
            return None
    
    def _obtener_valor_numerico(self, fila: Tuple, indice: int) -> Optional[float]:
        """Obtiene un valor numérico de una celda, manejando fórmulas si es necesario"""
        try:
            if indice >= len(fila):
                return None
                
            valor = fila[indice]
            
            # Si es None o vacío
            if valor is None or valor == "":
                return None
            
            # Si ya es un número
            if isinstance(valor, (int, float)):
                return float(valor)
            
            # Si es string, intentar convertir
            if isinstance(valor, str):
                valor_limpio = valor.strip()
                
                # Si es una fórmula de Excel sin calcular
                if valor_limpio.startswith('='):
                    print(f"⚠️ [FORMULA] Fórmula detectada: {valor_limpio}")
                    
                    # Intentar evaluar fórmulas simples como =+E1001*(1+F1001)
                    formula_calculada = self._evaluar_formula_simple(valor_limpio, fila)
                    if formula_calculada is not None:
                        print(f"✅ [FORMULA] Fórmula calculada: {valor_limpio} = {formula_calculada}")
                        return formula_calculada
                    else:
                        print(f"❌ [FORMULA] No se pudo calcular la fórmula: {valor_limpio}")
                        return None
                
                # Intentar conversión directa a número
                try:
                    return float(valor_limpio)
                except ValueError:
                    return None
            
            return None
        except Exception as e:
            print(f"❌ [ERROR] Error obteniendo valor numérico en índice {indice}: {e}")
            return None
    
    def _evaluar_formula_simple(self, formula: str, fila: Tuple) -> Optional[float]:
        """Evalúa fórmulas simples de Excel como =+E1001*(1+F1001)"""
        try:
            # Remover el signo = al inicio
            formula_sin_igual = formula[1:].strip()
            
            # Si empieza con +, removerlo también
            if formula_sin_igual.startswith('+'):
                formula_sin_igual = formula_sin_igual[1:].strip()
            
            # Para fórmulas como E1001*(1+F1001), necesitamos los valores de las columnas
            # Por ahora, vamos a detectar patrones comunes
            
            # Patrón: columna_E * (1 + columna_F) - fórmula típica de precio con ganancia
            import re
            patron_precio_ganancia = re.match(r'[A-Z]+\d+\*\(1\+[A-Z]+\d+\)', formula_sin_igual)
            
            if patron_precio_ganancia:
                # Es una fórmula de precio + ganancia
                # Asumiendo que es precio_costo * (1 + porcentaje_ganancia)
                precio_costo = self._obtener_valor_directo(fila, 4)  # Columna E (índice 4)
                porcentaje_ganancia = self._obtener_valor_directo(fila, 5)  # Columna F (índice 5)
                
                if precio_costo is not None and porcentaje_ganancia is not None:
                    # Convertir porcentaje a decimal si es necesario
                    if porcentaje_ganancia > 1:
                        porcentaje_ganancia = porcentaje_ganancia / 100
                    
                    resultado = precio_costo * (1 + porcentaje_ganancia)
                    return resultado
            
            return None
            
        except Exception as e:
            print(f"❌ [ERROR] Error evaluando fórmula {formula}: {e}")
            return None
    
    def _obtener_valor_directo(self, fila: Tuple, indice: int) -> Optional[float]:
        """Obtiene un valor directo sin procesamiento de fórmulas"""
        try:
            if indice >= len(fila):
                return None
                
            valor = fila[indice]
            
            if valor is None or valor == "":
                return None
            
            if isinstance(valor, (int, float)):
                return float(valor)
            
            if isinstance(valor, str):
                try:
                    return float(valor.strip())
                except ValueError:
                    return None
            
            return None
        except:
            return None
    
    def importar_productos_bd(self, confirmar_importacion: bool = True) -> Dict:
        """
        Importa los productos validados a la base de datos
        
        Args:
            confirmar_importacion: Si confirmar o cancelar importación
            
        Returns:
            Dict con resultado de la importación
        """
        if not confirmar_importacion:
            # Limpiar archivo temporal
            self._limpiar_archivo_temporal()
            return {
                'error': False,
                'mensaje': 'Importación cancelada por el usuario',
                'importados': 0
            }
        
        if not self.productos_validados:
            return {
                'error': True,
                'mensaje': 'No hay productos válidos para importar'
            }
        
        try:
            # Usar la conexión SQLAlchemy de Flask
            productos_importados = 0
            productos_con_error = []
            
            for i, producto_data in enumerate(self.productos_validados):
                try:
                    # Actualizar progreso
                    self.progreso_actual = int((i + 1) * 100 / len(self.productos_validados))
                    
                    # Obtener o crear proveedor
                    proveedor = self._obtener_o_crear_proveedor(producto_data['proveedor_nombre'])
                    
                    # Obtener o crear categoría (usar categoría por defecto)
                    categoria = self._obtener_o_crear_categoria("General")
                    
                    # Obtener o crear marca (usar marca por defecto)
                    marca = self._obtener_o_crear_marca("Sin Marca")
                    
                    # Crear nuevo producto
                    nuevo_producto = Producto(
                        nombre=producto_data['nombre'],
                        precio=producto_data['precio_calculado'],
                        precio_costo=producto_data['precio_costo'],
                        porcentaje_ganancia=producto_data['porcentaje_ganancia'],
                        precio_venta_publico=producto_data['precio_ganancia_paquete'],
                        precio_por_100gr=producto_data['precio_por_unidad_100gr'],
                        descripcion=producto_data['descripcion'],
                        disponible=True,
                        proveedor_id=proveedor.id,
                        categoria_id=categoria.id,
                        marca_id=marca.id
                    )
                    
                    # Calcular precios automáticamente
                    nuevo_producto.calcular_precios()
                    
                    # Agregar a la sesión
                    db.session.add(nuevo_producto)
                    productos_importados += 1
                    
                    # Commit cada 10 productos para no sobrecargar
                    if (i + 1) % 10 == 0:
                        db.session.commit()
                    
                except Exception as e:
                    productos_con_error.append({
                        'fila': producto_data['fila'],
                        'nombre': producto_data['nombre'],
                        'error': str(e)
                    })
                    # Rollback de la transacción actual
                    db.session.rollback()
                    continue
            
            # Commit final
            db.session.commit()
            
            # Limpiar archivo temporal
            self._limpiar_archivo_temporal()
            
            return {
                'error': False,
                'mensaje': f'Importación completada exitosamente',
                'productos_importados': productos_importados,
                'productos_con_error': len(productos_con_error),
                'errores_detalle': productos_con_error,
                'progreso_final': 100
            }
            
        except Exception as e:
            db.session.rollback()
            return {
                'error': True,
                'mensaje': f'Error durante la importación: {str(e)}',
                'detalle': traceback.format_exc()
            }
    
    def _obtener_o_crear_proveedor(self, nombre_proveedor: str) -> Proveedor:
        """Obtiene proveedor existente o crea uno nuevo"""
        proveedor = Proveedor.query.filter_by(nombre=nombre_proveedor).first()
        if not proveedor:
            proveedor = Proveedor(nombre=nombre_proveedor)
            db.session.add(proveedor)
            db.session.flush()  # Para obtener el ID sin commit
        return proveedor
    
    def _obtener_o_crear_categoria(self, nombre_categoria: str) -> Categoria:
        """Obtiene categoría existente o crea una nueva"""
        categoria = Categoria.query.filter_by(nombre=nombre_categoria).first()
        if not categoria:
            categoria = Categoria(nombre=nombre_categoria)
            db.session.add(categoria)
            db.session.flush()
        return categoria
    
    def _obtener_o_crear_marca(self, nombre_marca: str) -> Marca:
        """Obtiene marca existente o crea una nueva"""
        marca = Marca.query.filter_by(nombre=nombre_marca).first()
        if not marca:
            marca = Marca(nombre=nombre_marca)
            db.session.add(marca)
            db.session.flush()
        return marca
    
    def _limpiar_archivo_temporal(self):
        """Elimina el archivo temporal si existe"""
        if self.archivo_temporal and os.path.exists(self.archivo_temporal):
            try:
                os.unlink(self.archivo_temporal)
                self.archivo_temporal = None
            except:
                pass
    
    def obtener_progreso(self) -> Dict:
        """Retorna el progreso actual de la importación"""
        return {
            'progreso': self.progreso_actual,
            'mensaje': f"Importando productos ({self.progreso_actual}%)",
            'total_productos': len(self.productos_validados),
            'completado': self.progreso_actual >= 100
        }

# Instancia global del importador
importador_global = ImportadorProductos()

@importador_bp.route('/importar/subir-excel', methods=['POST'])
def subir_archivo_excel():
    """
    Endpoint para subir archivo Excel y obtener vista previa de productos
    
    Returns:
        JSON con productos válidos e inválidos para vista previa
    """
    try:
        print("🔍 [IMPORTADOR DEBUG] Iniciando subida de archivo Excel...")
        print(f"🔍 [DEBUG] Method: {request.method}")
        print(f"🔍 [DEBUG] Content-Type: {request.content_type}")
        print(f"🔍 [DEBUG] Content-Length: {request.content_length}")
        print(f"🔍 [DEBUG] Headers: {dict(request.headers)}")
        print(f"🔍 [DEBUG] Form data: {dict(request.form)}")
        print(f"🔍 [DEBUG] Files: {list(request.files.keys())}")
        
        # Verificar que se haya enviado un archivo
        if 'archivo' not in request.files:
            print("❌ [ERROR] No se encontró 'archivo' en request.files")
            return jsonify({
                'error': True,
                'mensaje': 'No se ha enviado ningún archivo',
                'debug_info': {
                    'files_received': list(request.files.keys()),
                    'form_data': dict(request.form),
                    'expected_key': 'archivo'
                }
            }), 400
        
        archivo = request.files['archivo']
        print(f"🔍 [DEBUG] Archivo recibido: {archivo.filename}")
        
        # Verificar que el archivo tenga un nombre
        if archivo.filename == '':
            print("❌ [ERROR] Archivo sin nombre")
            return jsonify({
                'error': True,
                'mensaje': 'No se ha seleccionado ningún archivo',
                'debug_info': {
                    'filename': archivo.filename,
                    'content_type': archivo.content_type
                }
            }), 400
        
        # Verificar extensión del archivo
        if not archivo.filename.lower().endswith(('.xlsx', '.xls')):
            print(f"❌ [ERROR] Extensión no válida: {archivo.filename}")
            return jsonify({
                'error': True,
                'mensaje': 'El archivo debe ser un Excel (.xlsx o .xls)',
                'debug_info': {
                    'filename': archivo.filename,
                    'valid_extensions': ['.xlsx', '.xls'],
                    'received_extension': os.path.splitext(archivo.filename)[1]
                }
            }), 400
        
        # Procesar archivo Excel
        print("📊 [IMPORTADOR] Procesando archivo Excel...")
        resultado = importador_global.procesar_archivo_excel(archivo)
        
        if resultado['error']:
            print(f"❌ [ERROR] Error procesando archivo: {resultado.get('mensaje', 'Error desconocido')}")
            return jsonify(resultado), 400
        
        print(f"✅ [IMPORTADOR] Archivo procesado exitosamente")
        print(f"📊 [STATS] Total procesadas: {resultado['total_procesadas']}")
        print(f"📊 [STATS] Válidas: {resultado['total_validas']}")
        print(f"📊 [STATS] Inválidas: {resultado['total_invalidas']}")
        
        return jsonify({
            'error': False,
            'mensaje': 'Archivo procesado exitosamente',
            'vista_previa': {
                'total_filas_procesadas': resultado['total_procesadas'],
                'productos_validos': {
                    'cantidad': resultado['total_validas'],
                    'lista': resultado['productos_validados']
                },
                'productos_invalidos': {
                    'cantidad': resultado['total_invalidas'],
                    'lista': resultado['productos_excluidos']
                }
            },
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        print(f"❌ [ERROR CRÍTICO] Error inesperado en subir_archivo_excel: {str(e)}")
        print(f"❌ [TRACEBACK] {traceback.format_exc()}")
        return jsonify({
            'error': True,
            'mensaje': f'Error inesperado: {str(e)}',
            'debug_info': {
                'error_type': type(e).__name__,
                'error_message': str(e),
                'traceback': traceback.format_exc(),
                'timestamp': datetime.now().isoformat()
            }
        }), 500

@importador_bp.route('/importar/confirmar', methods=['POST'])
def confirmar_importacion():
    """
    Endpoint para confirmar o cancelar la importación de productos
    
    Body JSON:
        {
            "confirmar": true/false
        }
    
    Returns:
        JSON con resultado de la importación
    """
    try:
        datos = request.get_json()
        
        if not datos:
            return jsonify({
                'error': True,
                'mensaje': 'Debe enviar datos JSON en el cuerpo de la petición'
            }), 400
        
        confirmar = datos.get('confirmar', False)
        
        # Ejecutar importación
        resultado = importador_global.importar_productos_bd(confirmar)
        
        if resultado['error']:
            return jsonify(resultado), 400
        
        return jsonify({
            'error': False,
            'resultado': resultado,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'error': True,
            'mensaje': f'Error inesperado: {str(e)}',
            'detalle': traceback.format_exc()
        }), 500

@importador_bp.route('/importar/progreso', methods=['GET'])
def obtener_progreso_importacion():
    """
    Endpoint para obtener el progreso actual de la importación
    
    Returns:
        JSON con progreso actual
    """
    try:
        progreso = importador_global.obtener_progreso()
        return jsonify({
            'error': False,
            'progreso': progreso,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'error': True,
            'mensaje': f'Error obteniendo progreso: {str(e)}'
        }), 500

@importador_bp.route('/importar/plantilla', methods=['GET'])
def descargar_plantilla_excel():
    """
    Endpoint para descargar plantilla de Excel con formato requerido
    
    Returns:
        Archivo Excel con la plantilla
    """
    try:
        # Crear un nuevo workbook
        workbook = openpyxl.Workbook()
        hoja = workbook.active
        hoja.title = "Plantilla Productos"
        
        # Definir encabezados según especificación
        encabezados = [
            "Producto",           # Columna A
            "",                   # Columna B (vacía)
            "Proveedor",          # Columna C  
            "",                   # Columna D (vacía)
            "Precio de Costo",    # Columna E
            "% de ganancia",      # Columna F
            "Precio de Ganancia del Paquete",  # Columna G
            "Precio por Unidad o 100 gramos",  # Columna H
            "Descripción"         # Columna I
        ]
        
        # Escribir encabezados en fila 1
        for col, encabezado in enumerate(encabezados, 1):
            hoja.cell(row=1, column=col, value=encabezado)
        
        # Agregar ejemplo en fila 2
        ejemplo = [
            "Arroz Integral 1kg",  # Producto
            "",                    # Vacío
            "Distribuidora ABC",   # Proveedor
            "",                    # Vacío
            25.50,                 # Precio de Costo
            40,                    # % ganancia
            35.70,                 # Precio Ganancia Paquete
            "",                    # Precio por 100gr (vacío = TIPO 1)
            "Arroz integral de alta calidad, rico en fibras"  # Descripción
        ]
        
        for col, valor in enumerate(ejemplo, 1):
            hoja.cell(row=2, column=col, value=valor)
        
        # Agregar otro ejemplo para TIPO 2
        ejemplo_tipo2 = [
            "Quinoa Premium",      # Producto
            "",                    # Vacío
            "Orgánicos del Sur",   # Proveedor
            "",                    # Vacío
            45.00,                 # Precio de Costo
            60,                    # % ganancia
            72.00,                 # Precio Ganancia Paquete
            7.20,                  # Precio por 100gr (con valor = TIPO 2)
            "Quinoa premium importada, sin gluten"  # Descripción
        ]
        
        for col, valor in enumerate(ejemplo_tipo2, 1):
            hoja.cell(row=3, column=col, value=valor)
        
        # Ajustar ancho de columnas
        anchos_columna = [20, 5, 18, 5, 15, 15, 25, 25, 30]
        for col, ancho in enumerate(anchos_columna, 1):
            hoja.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
        
        # Guardar en archivo temporal
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook.save(temp_file.name)
        workbook.close()
        
        # Leer archivo para enviarlo
        with open(temp_file.name, 'rb') as f:
            contenido_archivo = f.read()
        
        # Limpiar archivo temporal
        os.unlink(temp_file.name)
        
        from flask import Response
        return Response(
            contenido_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=plantilla_productos_deliciasnaturales.xlsx'
            }
        )
        
    except Exception as e:
        return jsonify({
            'error': True,
            'mensaje': f'Error generando plantilla: {str(e)}'
        }), 500

# Endpoint adicional para obtener estadísticas de importación
@importador_bp.route('/importar/estadisticas', methods=['GET'])
def obtener_estadisticas_importacion():
    """
    Obtiene estadísticas generales de productos en la BD
    """
    try:
        total_productos = Producto.query.count()
        total_proveedores = Proveedor.query.count()
        total_categorias = Categoria.query.count()
        total_marcas = Marca.query.count()
        
        # Productos por proveedor
        productos_por_proveedor = db.session.query(
            Proveedor.nombre,
            db.func.count(Producto.id).label('cantidad')
        ).join(Producto).group_by(Proveedor.id).all()
        
        return jsonify({
            'error': False,
            'estadisticas': {
                'totales': {
                    'productos': total_productos,
                    'proveedores': total_proveedores,
                    'categorias': total_categorias,
                    'marcas': total_marcas
                },
                'productos_por_proveedor': [
                    {'proveedor': p[0], 'cantidad': p[1]} 
                    for p in productos_por_proveedor
                ]
            },
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'error': True,
            'mensaje': f'Error obteniendo estadísticas: {str(e)}'
        }), 500
