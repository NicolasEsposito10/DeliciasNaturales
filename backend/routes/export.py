from flask import Blueprint, send_file, jsonify
from config.database import get_db
from datetime import datetime
import io
import csv

export_bp = Blueprint('export', __name__)

@export_bp.route('/export/excel')
def exportar_datos_excel():
    try:
        print("=== EXPORTANDO DATOS A EXCEL ===")
        
        # Conectar a la base de datos
        conn = get_db()
        cursor = conn.cursor()
        
        # Crear un archivo en memoria
        output = io.BytesIO()
        
        try:
            # Intentar usar openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            workbook = Workbook()
            # Remover la hoja por defecto
            workbook.remove(workbook.active)
            
            # Definir las tablas a exportar (excluyendo usuarios)
            tablas_exportar = [
                'producto', 'categoria', 'marca', 'proveedor', 'unidad', 
                'tipo_alimento', 'banner', 'imagen_producto', 'producto_etiquetas'
            ]
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for tabla in tablas_exportar:
                try:
                    print(f"Exportando tabla: {tabla}")
                    
                    # Verificar si la tabla existe
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (tabla,))
                    if not cursor.fetchone():
                        print(f"Tabla {tabla} no existe, saltando...")
                        continue
                    
                    # Obtener estructura de la tabla
                    cursor.execute(f"PRAGMA table_info({tabla})")
                    columnas_info = cursor.fetchall()
                    columnas = [col[1] for col in columnas_info]
                    
                    # Obtener datos
                    cursor.execute(f"SELECT * FROM {tabla}")
                    datos = cursor.fetchall()
                    
                    # Crear hoja de trabajo
                    ws = workbook.create_sheet(title=tabla.capitalize())
                    
                    # Escribir encabezados
                    for col_num, columna in enumerate(columnas, 1):
                        cell = ws.cell(row=1, column=col_num, value=columna)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Escribir datos
                    for row_num, fila in enumerate(datos, 2):
                        for col_num, valor in enumerate(fila, 1):
                            # Manejar datos BLOB (imágenes)
                            if isinstance(valor, bytes):
                                ws.cell(row=row_num, column=col_num, value="[IMAGEN_BLOB]")
                            else:
                                ws.cell(row=row_num, column=col_num, value=valor)
                    
                    # Ajustar ancho de columnas
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    print(f"Tabla {tabla} exportada: {len(datos)} filas")
                    
                except Exception as e:
                    print(f"Error exportando tabla {tabla}: {e}")
                    continue
            
            # Guardar el archivo en memoria
            workbook.save(output)
            output.seek(0)
            
            conn.close()
            
            # Generar nombre de archivo con fecha
            fecha = datetime.now().strftime('%Y-%m-%d_%H-%M')
            filename = f'delicias_naturales_{fecha}.xlsx'
            
            print(f"=== EXPORTACIÓN COMPLETADA: {filename} ===")
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except ImportError:
            # Si no está openpyxl, usar CSV como alternativa
            print("openpyxl no disponible, usando CSV...")
            
            # Crear ZIP con múltiples CSV
            import zipfile
            
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                tablas_exportar = [
                    'producto', 'categoria', 'marca', 'proveedor', 'unidad', 
                    'tipo_alimento', 'banner', 'imagen_producto', 'producto_etiquetas'
                ]
                
                for tabla in tablas_exportar:
                    try:
                        # Verificar si la tabla existe
                        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (tabla,))
                        if not cursor.fetchone():
                            continue
                        
                        # Obtener estructura y datos
                        cursor.execute(f"PRAGMA table_info({tabla})")
                        columnas_info = cursor.fetchall()
                        columnas = [col[1] for col in columnas_info]
                        
                        cursor.execute(f"SELECT * FROM {tabla}")
                        datos = cursor.fetchall()
                        
                        # Crear CSV en memoria
                        csv_buffer = io.StringIO()
                        writer = csv.writer(csv_buffer)
                        
                        # Escribir encabezados
                        writer.writerow(columnas)
                        
                        # Escribir datos
                        for fila in datos:
                            fila_procesada = []
                            for valor in fila:
                                if isinstance(valor, bytes):
                                    fila_procesada.append("[IMAGEN_BLOB]")
                                else:
                                    fila_procesada.append(valor)
                            writer.writerow(fila_procesada)
                        
                        # Agregar al ZIP
                        zip_file.writestr(f'{tabla}.csv', csv_buffer.getvalue())
                        
                    except Exception as e:
                        print(f"Error exportando tabla {tabla}: {e}")
                        continue
            
            zip_buffer.seek(0)
            conn.close()
            
            fecha = datetime.now().strftime('%Y-%m-%d_%H-%M')
            filename = f'delicias_naturales_{fecha}.zip'
            
            return send_file(
                zip_buffer,
                mimetype='application/zip',
                as_attachment=True,
                download_name=filename
            )
            
    except Exception as e:
        print(f"Error en exportación: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error exportando datos: {str(e)}'}), 500
